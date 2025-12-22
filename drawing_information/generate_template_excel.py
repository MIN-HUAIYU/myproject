#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
板式换热器业绩表 - 按照模板格式生成
"""

import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def extract_equipment_data(ocr_text):
    """从OCR文本中提取设备数据"""
    data = {
        "产品编号": "",
        "用户": "",
        "设备名称": "",
        "台数": 1,
        "单位重量": "",
        "热侧": "",
        "冷侧": "",
        "设计压力板": "",  # 板侧(热侧)压力
        "设计压力壳": "",  # 壳侧(冷侧)压力
        "设计温度板": "",  # 板侧(热侧)温度
        "设计温度壳": "",  # 壳侧(冷侧)温度
        "板片材质": "",    # 板片材质
        "设备型号": ""
    }

    # 1. 产品编号
    job_match = re.search(r'(?:JOB NO\.|产品编号)\s*\n\s*([A-Z0-9]+)', ocr_text)
    if job_match:
        data["产品编号"] = job_match.group(1).strip()

    # 2. 用户名称
    client_match = re.search(r'(?:业主|CLIENT)\s*(?:LCLIENT)?\s*(?:PROJECT NO\.)?\s*([^\n]+)', ocr_text)
    if client_match:
        client_text = client_match.group(1).strip()
        # 提取公司名称
        company_match = re.search(r'([^\s]+(?:公司|有限|股份|能源)[^\s]*)', client_text)
        if company_match:
            data["用户"] = company_match.group(1)

    # 3. 设备名称
    title_match = re.search(r'(?:DRAWING TITLE:|图纸名称)\s*[：:]*\s*\n\s*([^\n]+)', ocr_text)
    if title_match:
        data["设备名称"] = title_match.group(1).strip()

    # 4. 台数（默认1）
    data["台数"] = 1

    # 5. 单位重量 (kg)
    weight_match = re.search(r'设备净重\s*kg\s*(\d+)', ocr_text)
    if weight_match:
        data["单位重量"] = weight_match.group(1)

    # 6. 热侧介质
    if "低压蒸汽及凝液" in ocr_text:
        data["热侧"] = "低压蒸汽"
    elif "蒸汽" in ocr_text:
        data["热侧"] = "蒸汽"

    # 7. 冷侧介质
    if "碳酸钾溶液" in ocr_text:
        data["冷侧"] = "碳酸钾溶液"
    elif "溶液" in ocr_text:
        data["冷侧"] = "溶液"

    # 8. 设计压力 (板侧/壳侧)
    # 从"压力 设计 0.8/FV 1.0/FV"提取两个数值
    # 尝试多个格式
    pressure_patterns = [
        r'压力\s*设计\s*([\d.]+)/[A-Z]*\s+([\d.]+)/[A-Z]*',
        r'压力\s*设计\s*([\d.]+)\s+([\d.]+)',
        r'设计\s*([\d.]+)\s*/\s*(?:FV)?\s+([\d.]+)\s*/\s*(?:FV)?'
    ]
    for pattern in pressure_patterns:
        pressure_match = re.search(pattern, ocr_text)
        if pressure_match:
            data["设计压力板"] = pressure_match.group(1)  # 板侧(热侧)
            data["设计压力壳"] = pressure_match.group(2)  # 壳侧(冷侧)
            break

    # 9. 设计温度 (板侧/壳侧)
    # 从"温度 设计 220 220"提取两个数值
    temp_patterns = [
        r'温度\s*设计\s*(\d+)\s+(\d+)',
        r'温度\s*(?:℃)?\s*设计\s*(\d+)\s+(\d+)',
        r'设计\s*(\d+)\s*℃?\s+(\d+)\s*℃?'
    ]
    for pattern in temp_patterns:
        temp_match = re.search(pattern, ocr_text)
        if temp_match:
            data["设计温度板"] = temp_match.group(1)  # 板侧(热侧)
            data["设计温度壳"] = temp_match.group(2)  # 壳侧(冷侧)
            break

    # 10. 提取板片材质
    if "316L" in ocr_text:
        data["板片材质"] = "316L"
    elif "316" in ocr_text:
        data["板片材质"] = "316"
    elif "304" in ocr_text:
        data["板片材质"] = "304"

    # 11. 设备型号
    model_match = re.search(r'(LTB\d+-\d+[A-Z]?-\d+-[\d.]+-[\d.]+)', ocr_text)
    if model_match:
        data["设备型号"] = model_match.group(1).strip()

    return data


def create_template_excel(results_data):
    """按照模板格式创建Excel"""

    wb = Workbook()
    ws = wb.active
    ws.title = "业绩表"

    # 设置列宽
    columns_width = {
        'A': 8,    # 序号
        'B': 12,   # 产品编号
        'C': 18,   # 用户
        'D': 18,   # 设备名称
        'E': 8,    # 台数
        'F': 12,   # 单位重量 Kg
        'G': 18,   # 热侧/冷侧
        'H': 16,   # 设计压力MPa(板/壳)
        'I': 16,   # 设计温度℃(板/壳)
        'J': 12,   # 板片材质
        'K': 16    # 设备型号
    }

    for col, width in columns_width.items():
        ws.column_dimensions[col].width = width

    # 样式定义
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 第1行：标题
    ws.merge_cells('A1:K1')
    ws['A1'] = "板式换热器业绩表"
    ws['A1'].fill = title_fill
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 30

    # 第2行：表头
    headers = [
        ('A', '序号'),
        ('B', '产品编号'),
        ('C', '用户'),
        ('D', '设备名称'),
        ('E', '台数'),
        ('F', '单位重量\nKg'),
        ('G', '热侧/冷侧'),
        ('H', '设计压力MPa\n(板/壳)'),
        ('I', '设计温度℃\n(板/壳)'),
        ('J', '板片材质'),
        ('K', '设备型号')
    ]

    for col, header_text in headers:
        cell = ws[f'{col}2']
        cell.value = header_text
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align

    ws.row_dimensions[2].height = 25

    # 提取数据
    ocr_results = results_data.get("ocr_results", {})
    row_num = 3
    seq_num = 1

    for image_name, ocr_data in ocr_results.items():
        if "ocr_result" in ocr_data:
            equipment = extract_equipment_data(ocr_data["ocr_result"])

            # 填充数据
            cells_data = [
                ('A', seq_num),  # 序号
                ('B', equipment.get("产品编号", "")),
                ('C', equipment.get("用户", "")),
                ('D', equipment.get("设备名称", "")),
                ('E', equipment.get("台数", 1)),
                ('F', equipment.get("单位重量", "")),
                ('G', f"{equipment.get('热侧', '')}/{equipment.get('冷侧', '')}"),
                ('H', f"{equipment.get('设计压力板', '')}/{equipment.get('设计压力壳', '')}"),
                ('I', f"{equipment.get('设计温度板', '')}/{equipment.get('设计温度壳', '')}"),
                ('J', equipment.get("板片材质", "")),
                ('K', equipment.get("设备型号", ""))
            ]

            for col, value in cells_data:
                cell = ws[f'{col}{row_num}']
                cell.value = value
                cell.border = border

                # 序号列居中
                if col == 'A':
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

            ws.row_dimensions[row_num].height = 20
            row_num += 1
            seq_num += 1

    # 保存文件
    output_file = "heat_exchanger_performance.xlsx"
    wb.save(output_file)
    return output_file


def main():
    # 读取阿里云OCR结果
    try:
        with open("all_ocr_results.json", "r", encoding="utf-8") as f:
            results = json.load(f)
    except FileNotFoundError:
        print("[ERROR] 未找到 all_ocr_results.json 文件")
        return

    print("[START] 按照模板格式生成业绩表...")
    print("=" * 60)

    # 生成Excel
    output_file = create_template_excel(results)

    print(f"\n[SUCCESS] 业绩表已生成: {output_file}")
    print(f"\n表格格式：")
    print(f"  - 表头：按照标准模板")
    print(f"  - 列数：11列 (序号、产品编号、用户、设备名称、台数、重量、介质、压力、温度、型号等)")
    print(f"  - 行数：数据行数")
    print(f"\n文件位置：")
    print(f"  C:\\Users\\user\\Desktop\\drawing_information\\{output_file}")

    # 打印提取的数据示例
    ocr_results = results.get("ocr_results", {})
    print(f"\n提取的数据示例：")
    print("=" * 60)
    for image_name, ocr_data in ocr_results.items():
        if "ocr_result" in ocr_data:
            equipment = extract_equipment_data(ocr_data["ocr_result"])
            print(f"\n{image_name}:")
            for key, value in equipment.items():
                if value:
                    print(f"  {key}: {value}")


if __name__ == "__main__":
    main()
