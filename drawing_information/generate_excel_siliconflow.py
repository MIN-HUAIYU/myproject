#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
SiliconFlow OCR结果转Excel - 一上一下格式
"""

import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def extract_detailed_info(ocr_text, image_name):
    """从OCR文本中智能提取所有设备信息"""
    data = {
        "产品编号": "",
        "用户名称": "",
        "设备名称": "",
        "重量": "",
        "热侧介质": "",
        "冷侧介质": "",
        "设计压力": "",
        "设计温度": "",
        "设备型号": "",
        "板片材质": "",
        "设备位号": "",
        "装置名称": ""
    }

    # 提取产品编号
    job_patterns = [
        r'JOB NO\.\s*\n\s*([A-Z0-9]+)',
        r'产品编号\s*([A-Z0-9]+)',
        r'(25[A-Z]{2}\d{3})'
    ]
    for pattern in job_patterns:
        match = re.search(pattern, ocr_text)
        if match:
            data["产品编号"] = match.group(1).strip()
            break

    # 提取设备名称
    title_patterns = [
        r'DRAWING TITLE:\s*\n\s*([^\n]+)',
        r'图纸名称\s*[：:]\s*([^\n]+)',
        r'图纸名称\s*\n\s*([^\n]+)'
    ]
    for pattern in title_patterns:
        match = re.search(pattern, ocr_text)
        if match:
            title = match.group(1).strip()
            if title and not title.startswith('26'):
                data["设备名称"] = title
                break

    # 提取设备位号
    location_patterns = [
        r'(\d{5}-E\d{3}[A-Z]{2})',
        r'(\d{5}-[A-Z]\d{3}[A-Z]{2})'
    ]
    for pattern in location_patterns:
        match = re.search(pattern, ocr_text)
        if match:
            data["设备位号"] = match.group(1).strip()
            break

    # 提取设备型号
    model_patterns = [
        r'(LTB\d+-\d+[A-Z]?-\d+-[\d.]+-[\d.]+)',
        r'(LTB[\d\-A-Z.]+)'
    ]
    for pattern in model_patterns:
        match = re.search(pattern, ocr_text)
        if match:
            model = match.group(1).strip()
            if len(model) > 5:
                data["设备型号"] = model
                break

    # 提取业主/用户名称
    client_patterns = [
        r'业主\s*(?:CLIENT)?\s*(?:：|:)?\s*([^\s\n]+(?:\s+[^\n]+)?)',
        r'CLIENT\s*([^\n]+)',
        r'(伊泰伊犁能源有限公司|[^\s\n]+(?:公司|有限|股份))'
    ]
    for pattern in client_patterns:
        match = re.search(pattern, ocr_text)
        if match:
            client = match.group(1) if match.groups() else match.group(0)
            data["用户名称"] = client.strip()
            break

    # 提取项目名称作为装置名称
    project_patterns = [
        r'项目名称\s*(?:PROJECT)?\s*(?:：|:)?\s*([^\n]+)',
        r'PROJECT\s*([^\n]+)',
        r'(伊泰伊犁\d+万吨[^\n]+)'
    ]
    for pattern in project_patterns:
        match = re.search(pattern, ocr_text)
        if match:
            project = match.group(1) if match.groups() else match.group(0)
            if project and len(project) > 3:
                data["装置名称"] = project.strip()
                break

    # 提取板片材质
    if "316L" in ocr_text:
        data["板片材质"] = "316L"
    elif "316" in ocr_text:
        data["板片材质"] = "316"

    # 提取设计温度
    temp_patterns = [
        r'设计\s*\n\s*(\d+)\s*℃',
        r'温度\s*(?:℃)?\s*设计\s*(\d+)',
        r'(\d+)\s*℃'
    ]
    for pattern in temp_patterns:
        match = re.search(pattern, ocr_text)
        if match:
            data["设计温度"] = match.group(1) + " ℃"
            break

    # 提取设计压力
    pressure_patterns = [
        r'压力\s*(?:MPa)?\(G\)?\s*设计\s*([\d.]+)',
        r'设计\s*([\d.]+)\s*MPa',
        r'([\d.]+)\s*/?\s*(?:FV)?\s*MPa'
    ]
    for pattern in pressure_patterns:
        match = re.search(pattern, ocr_text)
        if match:
            data["设计压力"] = match.group(1) + " MPa"
            break

    # 提取重量
    weight_patterns = [
        r'设备净重\s*kg\s*(\d+)',
        r'(\d{4,})\s*kg'
    ]
    for pattern in weight_patterns:
        match = re.search(pattern, ocr_text)
        if match:
            weight = match.group(1)
            if int(weight) > 100:
                data["重量"] = weight + " kg"
                break

    # 提取热侧/冷侧介质
    medium_pattern = r'介质\s*名称\s*([^\n]+)\s+([^\n]+)'
    medium_match = re.search(medium_pattern, ocr_text)
    if medium_match:
        mediums = medium_match.groups()
        data["热侧介质"] = mediums[0].strip()
        if len(mediums) >= 2:
            data["冷侧介质"] = mediums[1].strip()
    else:
        # 备选方案
        if "低压蒸汽及凝液" in ocr_text:
            data["热侧介质"] = "低压蒸汽及凝液"
        if "碳酸钾溶液" in ocr_text:
            data["冷侧介质"] = "碳酸钾溶液及蒸气"

    return data


def create_vertical_excel(equipment_list):
    """创建竖向一上一下格式的Excel"""

    wb = Workbook()
    ws = wb.active
    ws.title = "设备数据表"

    # 设置列宽
    ws.column_dimensions['A'].width = 18

    # 样式定义
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    field_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    field_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 字段列表（按照用户要求的顺序）
    fields = [
        "产品编号",
        "用户名称",
        "设备名称",
        "重量",
        "热侧介质",
        "冷侧介质",
        "设计压力",
        "设计温度",
        "设备型号",
        "板片材质",
        "设备位号",
        "装置名称"
    ]

    # 第一列：字段名
    ws['A1'] = "字段名"
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].border = border
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # 为每个设备添加列
    for col_idx, equipment in enumerate(equipment_list, 2):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 22

        # 设备标题
        equipment_title = equipment.get("产品编号") or equipment.get("设备名称") or f"设备{col_idx-1}"
        ws[f'{col_letter}1'] = equipment_title
        ws[f'{col_letter}1'].fill = header_fill
        ws[f'{col_letter}1'].font = header_font
        ws[f'{col_letter}1'].border = border
        ws[f'{col_letter}1'].alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True
        )
        ws.row_dimensions[1].height = 25

        # 填充数据
        for row_idx, field in enumerate(fields, 2):
            # 字段名
            cell_field = ws[f'A{row_idx}']
            cell_field.value = field
            cell_field.fill = field_fill
            cell_field.font = field_font
            cell_field.border = border
            cell_field.alignment = Alignment(
                horizontal='left', vertical='center', wrap_text=True
            )

            # 字段值
            cell_value = ws[f'{col_letter}{row_idx}']
            cell_value.value = equipment.get(field, "")
            cell_value.border = border
            cell_value.alignment = Alignment(
                horizontal='left', vertical='top', wrap_text=True
            )

            ws.row_dimensions[row_idx].height = 22

    # 添加说明
    total_row = len(fields) + 3
    ws[f'A{total_row}'] = "说明"
    ws[f'A{total_row}'].font = Font(bold=True, size=10)

    ws[f'A{total_row + 1}'] = "生成时间：" + datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws[f'A{total_row + 2}'] = f"识别图纸数：{len(equipment_list)}"
    ws[f'A{total_row + 3}'] = "OCR供应商：SiliconFlow (DeepSeek-OCR)"
    ws[f'A{total_row + 4}'] = "数据格式：第一列为字段名，后续列为各设备数据"

    # 保存文件
    output_file = "设备数据表_SiliconFlow.xlsx"
    wb.save(output_file)
    return output_file


def main():
    # 读取JSON结果
    try:
        with open("siliconflow_all_results.json", "r", encoding="utf-8") as f:
            results = json.load(f)
    except FileNotFoundError:
        print("[ERROR] 未找到 siliconflow_all_results.json 文件")
        print("[HINT] 请先运行 batch_recognize_siliconflow.py")
        return

    print("[START] 开始提取设备信息并生成Excel...")
    print("=" * 60)

    # 提取所有设备信息
    equipment_list = []
    ocr_results = results.get("ocr_results", {})

    for image_name, ocr_data in ocr_results.items():
        if "ocr_result" in ocr_data and ocr_data.get("status") == "success":
            print(f"\n[Processing] 正在处理: {image_name}")
            equipment_info = extract_detailed_info(ocr_data["ocr_result"], image_name)
            equipment_list.append(equipment_info)

            # 打印提取的信息
            extracted_count = sum(1 for v in equipment_info.values() if v)
            print(f"  已提取 {extracted_count} 个字段的信息")

    print("\n" + "=" * 60)

    # 创建Excel文件
    if equipment_list:
        output_file = create_vertical_excel(equipment_list)
        print(f"[SUCCESS] Excel文件已生成: {output_file}")
        print(f"\n文件信息：")
        print(f"  [OK] 识别设备数：{len(equipment_list)}")
        print(f"  [OK] 字段数：12 (产品编号、用户名称、设备名称等)")
        print(f"  [OK] 格式：竖向一上一下（字段与数值对应）")
        print(f"\n文件位置：")
        print(f"  {output_file}")
    else:
        print("[ERROR] 没有成功提取任何设备信息")


if __name__ == "__main__":
    main()
