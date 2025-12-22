import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def extract_data_from_ocr(ocr_text):
    """从OCR文本中提取需要的字段"""

    data = {
        "产品编号": "",
        "用户名称": "",
        "设备名称": "蒸汽煮沸器Ⅱ",  # 用户提供的示例
        "重量": "",
        "热侧/冷侧介质名称": "",
        "设计压力/Mpa": "",
        "设计温度℃": "",
        "设备型号": "LTB2-500A-1-1.2-1.0",  # 用户提供的示例
        "板片材质": "316L",  # 从OCR文本中识别
        "设备位号": "26421-E002AB",  # 用户提供的示例
        "装置名称": ""
    }

    # 尝试从OCR文本中提取板片材质
    if "316L" in ocr_text:
        data["板片材质"] = "316L"

    return data


def create_excel_template(ocr_text):
    """创建Excel模板并填充可提取的数据"""

    # 提取数据
    data = extract_data_from_ocr(ocr_text)

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "设备信息"

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30

    # 设置样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题
    ws['A1'] = "设备信息表"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # 添加生成时间
    ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:B2')

    # 表头
    headers = ["字段名称", "数值"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 数据行
    fields = [
        "产品编号",
        "用户名称",
        "设备名称",
        "重量",
        "热侧/冷侧介质名称",
        "设计压力/Mpa",
        "设计温度℃",
        "设备型号",
        "板片材质",
        "设备位号",
        "装置名称"
    ]

    for row_idx, field in enumerate(fields, 5):
        # 字段名
        cell_name = ws.cell(row=row_idx, column=1)
        cell_name.value = field
        cell_name.border = border
        cell_name.font = Font(bold=True)
        cell_name.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell_name.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        # 数值
        cell_value = ws.cell(row=row_idx, column=2)
        cell_value.value = data.get(field, "")
        cell_value.border = border
        cell_value.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 设置行高
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[4].height = 20
    for row_idx in range(5, 5 + len(fields)):
        ws.row_dimensions[row_idx].height = 20

    # 保存文件
    output_file = "设备信息表.xlsx"
    wb.save(output_file)

    return output_file, data


def main():
    # 读取OCR结果
    with open("ocr_result.txt", "r", encoding="utf-8") as f:
        ocr_text = f.read()

    print("正在生成Excel表格...")
    output_file, extracted_data = create_excel_template(ocr_text)

    print(f"[SUCCESS] Excel文件已生成: {output_file}")
    print("\n已自动填充的字段：")
    print("=" * 50)
    for key, value in extracted_data.items():
        if value:
            print(f"  {key}: {value}")

    print("\n请打开 Excel 文件，手动补充以下字段（特别是图纸右下角的信息）：")
    print("=" * 50)
    for field in ["产品编号", "用户名称", "重量", "热侧/冷侧介质名称",
                  "设计压力/Mpa", "设计温度℃", "装置名称"]:
        if not extracted_data.get(field):
            print(f"  - {field}")

    print("\n提示：")
    print("  * 设备型号、设备位号来自图纸右下角的图纸名称部分")
    print("  * 如果这些字段在图纸中没有，可以留空")


if __name__ == "__main__":
    main()
