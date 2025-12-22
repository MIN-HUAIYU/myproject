#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
基于阿里云百炼OCR结果生成Excel表格
改进版本 - 更好的字段提取
"""

import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def extract_equipment_info_improved(ocr_text, image_name):
    """改进的设备信息提取"""
    data = {
        "产品编号": "",
        "用户名称": "",
        "设备名称": "",
        "重量": "",
        "热侧介质": "",
        "冷侧介质": "",
        "设计压力": "",
        "设计温度": "",
        "设备型号": "",
        "板片材质": "",
        "设备位号": "",
        "装置名称": ""
    }

    # 1. 提取产品编号 (JOB NO.)
    job_match = re.search(r'(?:JOB NO\.|产品编号)\s*\n\s*([A-Z0-9]+)', ocr_text)
    if job_match:
        data["产品编号"] = job_match.group(1).strip()

    # 2. 提取图纸名称/设备名称 (DRAWING TITLE:)
    title_match = re.search(r'(?:DRAWING TITLE:|图纸名称)\s*[：:]*\s*\n\s*([^\n]+)', ocr_text)
    if title_match:
        data["设备名称"] = title_match.group(1).strip()

    # 3. 提取设备位号 (格式: XXXXX-EXXXXX)
    location_match = re.search(r'(\d{5}-E\d{3}[A-Z]{2})', ocr_text)
    if location_match:
        data["设备位号"] = location_match.group(1).strip()

    # 4. 提取设备型号 (LTB开头)
    model_match = re.search(r'(LTB\d+-\d+[A-Z]?-\d+-[\d.]+-[\d.]+)', ocr_text)
    if model_match:
        data["设备型号"] = model_match.group(1).strip()

    # 5. 提取业主/客户名称 (CLIENT 或 业主)
    client_match = re.search(r'(?:业主|CLIENT)\s*(?:LCLIENT)?\s*(?:PROJECT NO\.)?\s*([^\n]+)', ocr_text)
    if client_match:
        client_text = client_match.group(1).strip()
        # 提取公司名称
        company_match = re.search(r'([^\s]+(?:公司|有限|股份|能源)[^\s]*)', client_text)
        if company_match:
            data["用户名称"] = company_match.group(1)

    # 6. 提取项目名称作为装置名称
    # 尝试多个模式
    project_patterns = [
        r'PROJECT\s*([^\n]+)',
        r'项目名称\s*(?:PROJECT)?\s*([^\n]+)',
        r'(伊泰伊犁\d+万吨[^\n]+)',
        r'项目\s*([^\n]+?)\s*(?:业主|CLIENT)'
    ]
    for pattern in project_patterns:
        project_match = re.search(pattern, ocr_text)
        if project_match:
            project_text = project_match.group(1).strip()
            if project_text and len(project_text) > 5 and not project_text.startswith('CLIENT'):
                data["装置名称"] = project_text
                break

    # 7. 提取板片材质
    if "316L" in ocr_text:
        data["板片材质"] = "316L"
    elif "316" in ocr_text:
        data["板片材质"] = "316"

    # 8. 提取设计温度 (℃)
    # 多个模式尝试
    temp_patterns = [
        r'温度\s*(?:℃)?\s*设计\s*(\d+)',
        r'设计\s*(\d+)\s*℃',
        r'温度.*?设计.*?(\d+)\s*℃'
    ]
    for pattern in temp_patterns:
        temp_match = re.search(pattern, ocr_text)
        if temp_match:
            data["设计温度"] = temp_match.group(1) + " ℃"
            break

    # 9. 提取设计压力 (MPa)
    # 多个模式尝试
    pressure_patterns = [
        r'压力\s*(?:MPa)?\(G\)?\s*设计\s*([\d.]+)',
        r'设计\s*([\d.]+)\s*/(?:FV)?',
        r'压力.*?设计.*?([\d.]+)\s*/'
    ]
    for pattern in pressure_patterns:
        pressure_match = re.search(pattern, ocr_text)
        if pressure_match:
            data["设计压力"] = pressure_match.group(1) + " MPa"
            break

    # 10. 提取重量 (设备净重 kg XXXX)
    weight_match = re.search(r'设备净重\s*kg\s*(\d+)', ocr_text)
    if weight_match:
        data["重量"] = weight_match.group(1) + " kg"

    # 11. 提取热侧和冷侧介质
    # 从"介质 名称 XXX XXX"提取
    medium_match = re.search(r'介质\s*名称\s*([^\n]+?)\s+([^\n]+?)(?:\s+毒性|$)', ocr_text)
    if medium_match:
        mediums = medium_match.groups()
        if len(mediums) >= 1:
            data["热侧介质"] = mediums[0].strip()
        if len(mediums) >= 2:
            data["冷侧介质"] = mediums[1].strip()

    return data


def create_vertical_excel_from_aliyun(results_data):
    """从阿里云结果创建竖向Excel"""

    wb = Workbook()
    ws = wb.active
    ws.title = "设备数据表"

    # 设置列宽
    ws.column_dimensions['A'].width = 18

    # 样式定义
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    field_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    field_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 字段列表（按用户要求的顺序）
    fields = [
        "产品编号",
        "用户名称",
        "设备名称",
        "重量",
        "热侧介质",
        "冷侧介质",
        "设计压力",
        "设计温度",
        "设备型号",
        "板片材质",
        "设备位号",
        "装置名称"
    ]

    # 第一列：字段名
    ws['A1'] = "字段名"
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].border = border
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # 提取所有设备信息
    equipment_list = []
    ocr_results = results_data.get("ocr_results", {})

    for image_name, ocr_data in ocr_results.items():
        if "ocr_result" in ocr_data:
            equipment_info = extract_equipment_info_improved(ocr_data["ocr_result"], image_name)
            equipment_list.append(equipment_info)

    # 为每个设备添加列
    for col_idx, equipment in enumerate(equipment_list, 2):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 22

        # 设备标题
        equipment_title = equipment.get("产品编号") or equipment.get("设备名称") or "设备"
        ws[f'{col_letter}1'] = equipment_title
        ws[f'{col_letter}1'].fill = header_fill
        ws[f'{col_letter}1'].font = header_font
        ws[f'{col_letter}1'].border = border
        ws[f'{col_letter}1'].alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True
        )
        ws.row_dimensions[1].height = 25

        # 填充数据
        for row_idx, field in enumerate(fields, 2):
            # 字段名
            cell_field = ws[f'A{row_idx}']
            cell_field.value = field
            cell_field.fill = field_fill
            cell_field.font = field_font
            cell_field.border = border
            cell_field.alignment = Alignment(
                horizontal='left', vertical='center', wrap_text=True
            )

            # 字段值
            cell_value = ws[f'{col_letter}{row_idx}']
            cell_value.value = equipment.get(field, "")
            cell_value.border = border
            cell_value.alignment = Alignment(
                horizontal='left', vertical='top', wrap_text=True
            )

            ws.row_dimensions[row_idx].height = 22

    # 添加说明
    total_row = len(fields) + 3
    ws[f'A{total_row}'] = "说明"
    ws[f'A{total_row}'].font = Font(bold=True, size=10)

    ws[f'A{total_row + 1}'] = "生成时间：" + datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws[f'A{total_row + 2}'] = f"识别图纸数：{len(equipment_list)}"
    ws[f'A{total_row + 3}'] = "OCR提供商：阿里云百炼 (Qwen-VL-OCR)"
    ws[f'A{total_row + 4}'] = "数据格式：第一列为字段名，后续列为各设备数据"

    # 保存文件
    output_file = "设备数据表_阿里云优化版.xlsx"
    wb.save(output_file)
    return output_file, len(equipment_list)


def main():
    # 读取阿里云OCR结果
    try:
        with open("all_ocr_results.json", "r", encoding="utf-8") as f:
            results = json.load(f)
    except FileNotFoundError:
        print("[ERROR] 未找到 all_ocr_results.json 文件")
        return

    print("[START] 基于阿里云百炼OCR结果生成Excel...")
    print("=" * 60)

    # 生成Excel
    output_file, equipment_count = create_vertical_excel_from_aliyun(results)

    print(f"\n[SUCCESS] Excel文件已生成: {output_file}")
    print(f"\n文件信息：")
    print(f"  识别设备数：{equipment_count}")
    print(f"  字段数：12")
    print(f"  格式：竖向一上一下（字段与数值对应）")
    print(f"\n文件位置：")
    print(f"  C:\\Users\\user\\Desktop\\drawing_information\\{output_file}")

    # 打印提取的设备信息
    ocr_results = results.get("ocr_results", {})
    for image_name, ocr_data in ocr_results.items():
        if "ocr_result" in ocr_data:
            print(f"\n[INFO] {image_name} 提取的信息：")
            equipment_info = extract_equipment_info_improved(ocr_data["ocr_result"], image_name)
            for key, value in equipment_info.items():
                if value:
                    print(f"  {key}: {value}")


if __name__ == "__main__":
    main()
