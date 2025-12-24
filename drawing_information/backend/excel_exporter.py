# -*- coding: utf-8 -*-
"""
Excel 导出模块 - 将 OCR 提取的信息写入 Excel 表格
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO


def extract_equipment_info(ocr_text):
    """从 OCR 文本中提取设备信息"""
    data = {
        "产品编号": "",
        "用户信息": "",
        "设备名称": "",
        "台数": "",
        "单台重量": "",
        "热侧/冷侧介质名称": "",
        "板程/壳程介质名称": "",
        "设计压力/MPa": "",
        "设计温度/℃": "",
        "设备型号": "",
        "板片材质": "",
        "换热面积/㎡": ""
    }

    # 提取产品编号 (JOB NO.)
    job_match = re.search(r'(?:JOB NO\.|产品编号)\s*\n\s*([A-Z0-9\-]+)', ocr_text)
    if job_match:
        data["产品编号"] = job_match.group(1).strip()

    # 提取设备名称 (DRAWING TITLE:)
    title_match = re.search(r'(?:DRAWING TITLE:|图纸名称)\s*[：:]*\s*\n\s*([^\n]+)', ocr_text)
    if title_match:
        data["设备名称"] = title_match.group(1).strip()

    # 提取用户信息 (CLIENT 或 业主)
    client_match = re.search(r'(?:业主|CLIENT)\s*(?:LCLIENT)?\s*(?:PROJECT NO\.)?\s*([^\n]+)', ocr_text)
    if client_match:
        data["用户信息"] = client_match.group(1).strip()

    # 提取设备型号 (LTB开头 或其他型号)
    model_match = re.search(r'(LTB[^\s\n]+|PHE[^\s\n]+|[A-Z]+\d+[^\s\n]*)', ocr_text)
    if model_match:
        data["设备型号"] = model_match.group(1).strip()

    # 提取板片材质
    material_patterns = [r'316L', r'316', r'钛', r'铜镍', r'不锈钢']
    for pattern in material_patterns:
        if re.search(pattern, ocr_text):
            data["板片材质"] = pattern
            break

    # 提取台数
    qty_match = re.search(r'台数\s*[:：]?\s*(\d+)', ocr_text)
    if qty_match:
        data["台数"] = qty_match.group(1)
    else:
        # 尝试其他模式
        qty_match = re.search(r'(?:QTY|Qty)\s*[:：]?\s*(\d+)', ocr_text, re.IGNORECASE)
        if qty_match:
            data["台数"] = qty_match.group(1)

    # 提取单台重量 (kg)
    weight_match = re.search(r'(?:重量|Weight|单台重量)\s*[:：]?\s*(\d+\.?\d*)\s*(?:kg|Kg)', ocr_text)
    if weight_match:
        data["单台重量"] = weight_match.group(1) + " kg"

    # 提取设计温度 (℃)
    temp_patterns = [
        r'设计\s*(?:温度)?\s*(\d+)\s*℃',
        r'温度\s*[:：]?\s*(\d+)\s*℃',
        r'(?:Design\s+)?[Tt]emperature\s*[:：]?\s*(\d+)\s*°C',
    ]
    for pattern in temp_patterns:
        temp_match = re.search(pattern, ocr_text)
        if temp_match:
            data["设计温度/℃"] = temp_match.group(1)
            break

    # 提取设计压力 (MPa)
    pressure_patterns = [
        r'设计\s*(?:压力)?\s*([\d.]+)\s*(?:MPa|Mpa)',
        r'压力\s*[:：]?\s*([\d.]+)\s*(?:MPa|Mpa)',
        r'(?:Design\s+)?[Pp]ressure\s*[:：]?\s*([\d.]+)\s*(?:MPa|bar)',
    ]
    for pattern in pressure_patterns:
        pressure_match = re.search(pattern, ocr_text)
        if pressure_match:
            data["设计压力/MPa"] = pressure_match.group(1)
            break

    # 提取介质信息
    medium_patterns = [
        r'介质\s*名称\s*([^\n]+)',
        r'(?:热侧|冷侧)\s*[:：]?\s*([^\n]+)',
        r'[Mm]edium\s*[:：]?\s*([^\n]+)',
    ]
    for pattern in medium_patterns:
        medium_match = re.search(pattern, ocr_text)
        if medium_match:
            medium_text = medium_match.group(1).strip()
            # 尝试分离热侧和冷侧
            if '/' in medium_text:
                parts = medium_text.split('/')
                data["热侧/冷侧介质名称"] = parts[0].strip()
                if len(parts) > 1:
                    data["板程/壳程介质名称"] = parts[1].strip()
            else:
                data["热侧/冷侧介质名称"] = medium_text
            break

    # 提取换热面积
    area_match = re.search(r'(?:换热面积|面积|Heat[_\s]?Area)\s*[:：]?\s*([\d.]+)\s*(?:㎡|m²|m2)', ocr_text)
    if area_match:
        data["换热面积/㎡"] = area_match.group(1)

    return data


def create_horizontal_excel(extracted_data_list):
    """
    创建横排 Excel 表格
    Args:
        extracted_data_list: 提取的设备信息列表，每个元素是一个字典
    Returns:
        BytesIO: Excel 文件的字节流
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "设备数据表"

    # 定义表头
    headers = [
        "产品编号",
        "用户信息",
        "设备名称",
        "台数",
        "单台重量",
        "热侧/冷侧介质名称",
        "板程/壳程介质名称",
        "设计压力/MPa",
        "设计温度/℃",
        "设备型号",
        "板片材质",
        "换热面积/㎡"
    ]

    # 设置列宽
    col_widths = [15, 18, 18, 8, 12, 15, 15, 12, 12, 15, 12, 12]
    for col_idx, width in enumerate(col_widths, 1):
        col_letter = chr(64 + col_idx)  # A=65, B=66...
        ws.column_dimensions[col_letter].width = width

    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        col_letter = chr(64 + col_idx)
        cell = ws[f'{col_letter}1']
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 写入数据
    for row_idx, data in enumerate(extracted_data_list, 2):
        for col_idx, header in enumerate(headers, 1):
            col_letter = chr(64 + col_idx)
            cell = ws[f'{col_letter}{row_idx}']
            cell.value = data.get(header, "")
            cell.border = border
            cell.alignment = data_alignment

    # 保存到字节流
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_to_excel(ocr_result_text):
    """
    从 OCR 结果导出 Excel
    Args:
        ocr_result_text: OCR 识别的文本
    Returns:
        BytesIO: Excel 文件的字节流
    """
    # 提取设备信息
    equipment_data = extract_equipment_info(ocr_result_text)

    # 创建 Excel（支持单个或多个设备）
    data_list = [equipment_data]

    # 生成 Excel
    excel_file = create_horizontal_excel(data_list)

    return excel_file
